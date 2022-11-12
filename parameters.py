import openpyxl
from tkinter import *
from math import ceil as ceil
from imutils.perspective import four_point_transform
from imutils import contours
import numpy as np
import argparse
import imutils
import cv2
import os

def correct_cart(file, ANSWER_KEY, nroQuestions, vCorrect, vIncorrect, vBlank, fileName):
    image = cv2.imread(file)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 75, 200)

    cnts = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL,
        cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    docCnt = None

    if len(cnts) > 0:
        cnts = sorted(cnts, key=cv2.contourArea, reverse=True)
        for c in cnts:

            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)

            if len(approx) == 4:
                docCnt = approx
                break
                    
    paper = four_point_transform(image, docCnt.reshape(4, 2))
    warped = four_point_transform(gray, docCnt.reshape(4, 2))

    thresh = cv2.threshold(warped, 0, 255,
        cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL,
        cv2.CHAIN_APPROX_SIMPLE)
    cnts = imutils.grab_contours(cnts)
    questionCnts = []

    for c in cnts:
        (x, y, w, h) = cv2.boundingRect(c)
        ar = w / float(h)
        if w >= 20 and h >= 20 and ar >= 0.9 and ar <= 1.1:
            questionCnts.append(c)

    questionCnts = contours.sort_contours(questionCnts,
        method="top-to-bottom")[0]
    correct = 0

    for (q, i) in enumerate(np.arange(0, len(questionCnts), 5)):
        cnts = contours.sort_contours(questionCnts[i:i + 5])[0]
        bubbled = None

        for (j, c) in enumerate(cnts):
            mask = np.zeros(thresh.shape, dtype="uint8")
            cv2.drawContours(mask, [c], -1, 255, -1)


            mask = cv2.bitwise_and(thresh, thresh, mask=mask)
            total = cv2.countNonZero(mask)

            if bubbled is None or total > bubbled[0]:
                bubbled = (total, j)

        color = (0, 0, 255)
        k = ANSWER_KEY[q]

        if k == bubbled[1]:
            color = (0, 255, 0)
            correct += 1

        cv2.drawContours(paper, [cnts[k]], -1, color, 3)

    score = correct * vCorrect - (nroQuestions - correct) * vIncorrect - vBlank * (nroQuestions - correct - vCorrect)
    if score < 0:
        score = 0
    print("[INFO] score: {:.2f}".format(score))
    cv2.putText(paper, "Puntaje: {:.2f}".format(score), (10, 30),
        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
    save_path = "corregidas/rv_" + fileName
    cv2.imwrite(save_path,paper)
    # cv2.imshow("Correction", paper)
    # cv2.waitKey(0)


    return score

def close_window(root):
    root.destroy()

def parameters(root):
    root.destroy()
    param = Tk()
    param.title("Corrector de Examenes")
    param.geometry("1000x750")
    param.configure(background='white')


    #Labels
    Label(param, text="Parámetros del examen:", font=("Futura Md BT", 30),bg="white").grid(row=0, column=0,padx=30, pady=20,sticky=W)

    Label(param, text="¿Cuántas preguntas tiene el examen?", font=("Futura Bk BT", 23),bg="white").grid(row=1, column=0,padx=60,  pady=20,sticky=W)

    Label(param, text="¿Cuántos puntos vale una\nrespuesta correcta?", font=("Futura Bk BT", 23),bg="white").grid(row=3, column=0,padx=60, pady=20,sticky=W)

    Label(param, text="¿Cuántos puntos disminuye una\nrespuesta incorrecta?", font=("Futura Bk BT", 23),bg="white").grid(row=4,padx=60, column=0, pady=20,sticky=W)
    Label(param, text="¿Cuántos puntos disminuye una\nrespuesta en blanco?", font=("Futura Bk BT", 23),bg="white").grid(row=5,padx=60, column=0, pady=20,sticky=W)

    #Entrys
    txtQuestions = Entry(param, width=5, bg="#ECECEC", font= ("Futura Bk BT", 23))
    txtQuestions.grid(row=1, column=1, pady=20,sticky=W)
    txtCorrect = Entry(param, width=5, bg="#ECECEC", font= ("Futura Bk BT", 23))
    txtCorrect.grid(row=3, column=1, pady=20,sticky=W)
    txtIncorrect = Entry(param, width=5, bg="#ECECEC", font= ("Futura Bk BT", 23))
    txtIncorrect.grid(row=4, column=1, pady=20,sticky=W)
    txtBlank = Entry(param, width=5, bg="#ECECEC", font= ("Futura Bk BT", 23))
    txtBlank.grid(row=5, column=1, pady=20,sticky=W)

    def puntaje_window():
        puntaje = Toplevel()
        puntaje.title("Corrector de Examenes")
        puntaje.configure(background='white')

        #Privados
        col_answer = 0
        last_col = 0
        last_row = 0
        correct = []
        answers = []

        def correct_exam():
            ANSWER_KEY = {}

            for i in range(0, nroQuestions):
                ANSWER_KEY[i] = ord((answers[i].get()).upper())-65

            input_images_path = "cartillas"
            files_names = os.listdir(input_images_path)

            wb = openpyxl.Workbook()
            sheet = wb.active

            sheet_title = "Puntajes"

            sheet.cell(row=1, column=1).value = "Identificador"
            sheet.cell(row=1, column=2).value = "Puntaje"

            row = 2

            for file_name in files_names:
                image_path = input_images_path + "/" + file_name
                score = correct_cart(image_path, ANSWER_KEY, nroQuestions, vCorrect, vIncorrect, vBlank, file_name)
                sheet.cell(row=row, column=1).value = file_name[:-4]
                sheet.cell(row=row, column=2).value = score
                row += 1
            
            wb.save(sheet_title + ".xlsx")

        def set_questions():
            for i in range(1, nroQuestions + 1):
                Label(puntaje, text="Pregunta " + str(i) + ": ", bg="white").grid(row=(i-1)%25, column=(ceil(i/25)-1)*2)
                ans = Entry(puntaje, width=2, bg="#ECECEC")
                ans.grid(row=(i-1)%25,column= (ceil(i/25)-1)*2+1)
                answers.append(ans)
                last_row, last_col = (i-1)%25, (ceil(i/25)-1)*2+1

            Button(puntaje, text="Cerrar", command= puntaje.destroy).grid(row=last_row+1, column=last_col+1, sticky=W+E)
            Button(puntaje, text="Corregir", command=lambda : correct_exam(), bg="#42B4FF", fg="white").grid(row=last_row+1, column=last_col+2, sticky=W+E)

        #Parametros
        nroQuestions = int(txtQuestions.get())
        set_questions()
        vCorrect = float(txtCorrect.get())
        vIncorrect = float(txtIncorrect.get())
        vBlank = float(txtBlank.get())

    #Buttons
    Button(param, text="Siguiente", font=("Futura Bk BT", 23), bg="#42B4FF", fg="white", command=puntaje_window).grid(row=6, column=2, pady=20,sticky=W)
    Button(param, text="Salir", font=("Futura Bk BT", 23), bg="white", fg="black", command=lambda: close_window(param)).grid(row=6, column=1, pady=20,sticky=W)
